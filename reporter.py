import logging
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

logger = logging.getLogger(__name__)

COLOR_HEADER_BG    = "1F3864"   # Dark navy
COLOR_HEADER_FONT  = "FFFFFF"
COLOR_BUDGET       = "C6EFCE"   # Light green
COLOR_MID          = "FFEB9C"   # Light yellow
COLOR_PREMIUM      = "FFC7CE"   # Light red/pink
COLOR_DISCOUNT_ROW = "EBF5FB"   # Light blue for discounted rows
COLOR_SUBHEADER    = "D6E4F0"

def header_style(ws, row: int, columns: list[str]):
    fill   = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    font   = Font(color=COLOR_HEADER_FONT, bold=True, size=11)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_letter in columns:
        cell = ws[f"{col_letter}{row}"]
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border

def apply_segment_color(ws, row: int, segment: str, n_cols: int):
    color_map = {
        "Budget":    COLOR_BUDGET,
        "Mid-range": COLOR_MID,
        "Premium":   COLOR_PREMIUM,
    }
    color = color_map.get(segment, "FFFFFF")
    fill  = PatternFill("solid", fgColor=color)
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill

def write_summary_sheet(wb, summary:dict, keywords: list[str]):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24

    title_font = Font(bold=True, size=14, color=COLOR_HEADER_BG)
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    section_fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)

    def write_row(r, label, value, bold_val=False):
        ws.cell(r, 1, label).font = label_font
        cell = ws.cell(r, 2, value)
        cell.font = Font(bold=bold_val, size=11)
        cell.alignment = Alignment(horizontal="right")

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "Competitor Analysis Report"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Metadata
    ws.cell(3, 1, "Generated At").font = label_font
    ws.cell(3, 2, datetime.now().strftime("%Y-%m-%d %H:%M")).font = value_font
    ws.cell(4, 1, "Keywords Analyzed").font = label_font
    ws.cell(4, 2, ", ".join(keywords)).font = value_font

    # Stats section
    ws.merge_cells("A6:B6")
    ws["A6"] = "Key Metrics"
    ws["A6"].font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = Alignment(horizontal="center")

    write_row(7, "Total Products Scraped", summary.get("total_products", 0), bold_val=True)
    write_row(8, "Min Price", f"Rp {summary.get('price_min', 0):,.0f}")
    write_row(9, "Max Price", f"Rp {summary.get('price_max', 0):,.0f}")
    write_row(10, "Average Price", f"Rp {summary.get('price_mean', 0):,.0f}")
    write_row(11, "Median Price", f"Rp {summary.get('price_median', 0):,.0f}")
    write_row(12, "Average Rating", f"{summary.get('avg_rating', 0):.2f} / 5.0")
    write_row(13, "Products With Discount", f"{summary.get('pct_with_discount', 0):.1f}%")
    write_row(14, "Avg Discount %", f"{summary.get('avg_discount_pct', 0):.1f}%")

    # Segment distribution
    ws.merge_cells("A16:B16")
    ws["A16"] = "Price Segment Distribution"
    ws["A16"].font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
    ws["A16"].fill = section_fill
    ws["A16"].alignment = Alignment(horizontal="center")

    dist = summary.get("price_segment_distribution", {})
    total = sum(dist.values()) or 1
    r = 17
    for segment, count in dist.items():
        write_row(r, segment, f"{count} products ({count/total*100:.1f}%)")
        r += 1

    ws.merge_cells(f"A{r + 1}:B{r + 1}")
    ws[f"A{r + 1}"] = "Top Seller Locations"
    ws[f"A{r + 1}"].font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
    ws[f"A{r + 1}"].fill = section_fill
    ws[f"A{r + 1}"].alignment = Alignment(horizontal="center")

    r += 2
    for loc, count in summary.get("top_locations", {}).items():
        write_row(r, loc or "Unknown", f"{count} sellers")
        r += 1

    logger.info("Summary sheet written.")

def write_data_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Product Data")

    display_columns = [
        "keyword", "name", "price", "original_price",
        "discount_percent", "rating", "sold_count",
        "shop_name", "shop_location", "price_segment",
        "has_discount", "effective_discount_value", "url", "scraped_at"
    ]
    display_headers = [
        "Keyword", "Product Name", "Price (Rp)", "Original Price (Rp)",
        "Discount %", "Rating", "Sold Count",
        "Shop Name", "Location", "Price Segment",
        "Has Discount", "Discount Value (Rp)", "URL", "Scraped At"
    ]

    existing = [c for c in display_columns if c in df.columns]
    header_labels = [display_headers[display_columns.index(c)] for c in existing]

    col_letters = [get_column_letter(i+1) for i in range(len(existing))]

    for i, (col, label) in enumerate(zip(col_letters, header_labels), start=1):
        ws.cell(1, i, label)

    header_style(ws, 1, col_letters)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, (_, row) in enumerate(df[existing].iterrows(), start=2):
        for col_idx, col_name in enumerate(existing, start=1):
            value = row[col_name]
            # Convert numpy bools
            if hasattr(value, "item"):
                value = value.item()
            ws.cell(row_idx, col_idx, value)

        # Row color by segment
        segment = row.get("price_segment", "")
        apply_segment_color(ws, row_idx, segment, len(existing))

    # Format price columns
    accounting_fmt = '_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"_);_(@_)'
    price_cols = {"price", "original_price", "effective_discount_value"}
    for i, col_name in enumerate(existing, start=1):
        col_letter = get_column_letter(i)
        if col_name in price_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = accounting_fmt

    # Auto-fit column widths
    col_widths = {
        "name": 45, "url": 20, "shop_name": 22, "keyword": 18,
        "scraped_at": 18, "shop_location": 18,
    }
    for i, col_name in enumerate(existing, start=1):
        letter = get_column_letter(i)
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[letter].width = width

    logger.info(f"Product data sheet written: {len(df)} rows.")
    return ws

def add_price_chart(wb, df: pd.DataFrame):
    if df.empty or "keyword" not in df.columns:
        return

    agg = df.groupby("keyword")["price"].mean().reset_index()
    ws = wb.create_sheet("Price Chart")

    ws["A1"] = "Keyword"
    ws["B1"] = "Avg Price (Rp)"
    for i, (_, row) in enumerate(agg.iterrows(), start=2):
        ws[f"A{i}"] = row["keyword"]
        ws[f"B{i}"] = round(row["price"])

    accounting_fmt = '_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"_);_(@_)'
    for i in range(2, len(agg) + 2):
        ws[f"B{i}"].number_format = accounting_fmt

    chart = BarChart()
    chart.type  = "col"
    chart.title = "Average Price by Keyword"
    chart.y_axis.title = "Price (Rp)"
    chart.x_axis.title = "Keyword"
    chart.style = 10
    chart.width  = 20
    chart.height = 12

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(agg)+1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(agg)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    series = chart.series[0]
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True       # tampilkan nilai
    series.dLbls.showLegendKey = False
    series.dLbls.showCatName = False
    series.dLbls.showSerName = False
    series.dLbls.showPercent = False

    ws.add_chart(chart, "D2")
    logger.info("Price chart sheet added.")

def generate_excel_report(df: pd.DataFrame, summary: dict, keywords: list[str], output_path: str = "report.xlsx"):
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_summary_sheet(wb, summary, keywords)
    write_data_sheet(wb, df)
    add_price_chart(wb, df)

    wb.save(output_path)
    logger.info(f"Excel report saved to: {output_path}")
    return output_path